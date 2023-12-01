import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from _config import Config
from openpyxl.styles import NamedStyle, Font, Border, Side
import openpyxl.utils.cell as pyxlcell

def generate_rejected_excel(rejected_cases:pd.DataFrame, rejected_blueprint:pd.DataFrame, config:Config):
    rejected_cases = _format_dataframe(rejected_cases)
    filename = "Rejected_Report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    
    [sheet.append(row) for row in dataframe_to_rows(rejected_cases, index= False, header= True)]
    headers = {cell.value: cell for cell in sheet[1]}
    col_styles = _gen_col_styles()
    
    # def assignStyle(cell, cstyle): cell.style = cstyle
    
    autoHeaders = config.data['configuration']['importCSV']['autoHeaders']
    iheaders = config.data['configuration']['importCSV']['headers']
    if autoHeaders == ('CLEAN' or 'AUTO'):
        # The Saif Way
        # [[assignStyle(c, col_styles[iheaders[h]['type']]) for c in sheet[headers[h].column]] for h in headers.keys() & iheaders.keys()]
        i = 0
        for h in headers.keys() & iheaders.keys():
            for c in sheet[pyxlcell.get_column_letter(headers[h].column)]:
                c.style = col_styles[iheaders[h]['type']]

    
    workbook.save(filename)
    print("Rejected Workbook Saved.")
    return


def generate_accepted_excel(accepted_batch, config:Config):
    accepted_batch = _format_dataframe(accepted_batch)
    
    filename = "Accepted_Report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    
    for row in dataframe_to_rows(accepted_batch, index= False, header= True):
        sheet.append(row)
    workbook.save(filename)
    print("Accepted Workbook Saved.")
    return


def _format_dataframe(dataFrame):
    df = dataFrame.map(lambda x: None if (type(x) == str and x == 'None') else x)
    df = df.reset_index()
    return df

def _gen_col_styles():
    # Create styles here.
    
    
    
    style_dict = {"Currency" : "Currency", 
                  "Percent" : "Percent",
                  "Decimal" : "Normal",
                  "String" : "Normal",
                  "Integer" : "Normal",
                  "DateTime" : "Normal",
                  "Boolean" : "Normal"}
    
    return style_dict
# Function which creates a dictionary of styles.
    # Currency'ise style
    # Convert date-time 
    # Different highlight colours
    